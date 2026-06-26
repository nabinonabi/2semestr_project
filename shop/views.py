from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
import openpyxl
from django.core.mail import EmailMessage
from django.conf import settings
from io import BytesIO
from django.shortcuts import render

from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.models import User

from rest_framework.viewsets import ModelViewSet
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated, BasePermission, SAFE_METHODS
from rest_framework.response import Response
from rest_framework import status

from .models import Product, Category, Proizvodstvo, Cart, CartItem, Order, OrderItem, Profile
from .serializers import (
    ProductSerializer, CategorySerializer, ProizvodstvoSerializer,
    CartSerializer, CartItemSerializer, OrderSerializer, ProfileSerializer
)


def index(request):
    popular_products = Product.objects.all()[:8]
    categories = Category.objects.all()
    context = {
        'popular_products': popular_products,
        'categories': categories,
    }
    return render(request, 'shop/index.html', context)


def author_page(request):
    content = """
    <p>Автор: Кучина Маргарита, группа 89 ТП</p>
    <hr>
    <a href='/'><button>На главную</button></a>
    """
    return HttpResponse(content)


def shop_info_page(request):
    content = """
    <p>Это магазин товаров для кемпинга! Здесь скоро будут спальные мешки и палатки)</p>
    <hr>
    <a href='/'><button>На главную</button></a>
    """
    return HttpResponse(content)


def product_list(request):
    products = Product.objects.all()
    
    category_id = request.GET.get('category')
    if category_id:
        products = products.filter(category_id=category_id)
    
    proizvodstvo_id = request.GET.get('proizvodstvo')
    if proizvodstvo_id:
        products = products.filter(proizvodstvo_id=proizvodstvo_id)
    
    search_query = request.GET.get('search')
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) | 
            Q(description__icontains=search_query)
        )
    
    paginator = Paginator(products, 9)
    page_number = request.GET.get('page')
    products_page = paginator.get_page(page_number)
    
    categories = Category.objects.all()
    proizvodstvos = Proizvodstvo.objects.all()
    
    context = {
        'products': products_page,
        'categories': categories,
        'proizvodstvos': proizvodstvos,
        'selected_category': int(category_id) if category_id else None,
        'selected_proizvodstvo': int(proizvodstvo_id) if proizvodstvo_id else None,
        'search_query': search_query,
    }
    return render(request, 'shop/product_list.html', context)


def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    return render(request, 'shop/product_detail.html', {'product': product})


@login_required
def cart_add(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    
    if product.stock_quantity <= 0:
        messages.error(request, f'Товар "{product.name}" отсутствует на складе')
        return redirect('shop:product_detail', pk=product_id)
    
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart_item, item_created = CartItem.objects.get_or_create(
        cart=cart,
        product=product,
        defaults={'quantity': 1}
    )
    
    if not item_created:
        try:
            if cart_item.quantity < product.stock_quantity:
                cart_item.quantity += 1
                cart_item.save()
                messages.success(request, f'Количество товара "{product.name}" увеличено до {cart_item.quantity}')
            else:
                messages.error(request, f'Нельзя добавить больше {product.stock_quantity} шт. товара "{product.name}"')
        except ValidationError as e:
            messages.error(request, str(e))
    else:
        messages.success(request, f'Товар "{product.name}" добавлен в корзину')
    
    return redirect('shop:cart_view')


@login_required
def update_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    product = cart_item.product
    
    if request.method == 'POST':
        try:
            new_quantity = int(request.POST.get('quantity', 1))
            if new_quantity <= 0:
                cart_item.delete()
                messages.success(request, f'Товар "{product.name}" удален из корзины')
            else:
                if new_quantity > product.stock_quantity:
                    messages.warning(request, f'Максимум доступно: {product.stock_quantity} шт.')
                    new_quantity = product.stock_quantity
                
                cart_item.quantity = new_quantity
                cart_item.save() 
                messages.success(request, f'Количество товара "{product.name}" обновлено')
        except ValidationError as e:
            messages.error(request, str(e))
        except ValueError:
            messages.error(request, 'Некорректное значение количества')
    
    return redirect('shop:cart_view')


@login_required
def remove_from_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    product_name = cart_item.product.name
    cart_item.delete()
    messages.success(request, f'Товар "{product_name}" удален из корзины')
    return redirect('shop:cart_view')


@login_required
def cart_view(request):
    try:
        cart = Cart.objects.get(user=request.user)
        cart_items = cart.items.select_related('product').all()
        for item in cart_items:
            item.total_price = item.product.price * item.quantity
        total_price = cart.total_price()
    except Cart.DoesNotExist:
        cart_items = []
        total_price = 0
    
    context = {
        'cart_items': cart_items,
        'total_price': total_price,
    }
    return render(request, 'shop/cart.html', context)


@login_required
def checkout(request):
    try:
        cart = Cart.objects.get(user=request.user)
        cart_items = cart.items.select_related('product').all()
        
        if not cart_items:
            messages.warning(request, 'Ваша корзина пуста. Добавьте товары перед оформлением заказа.')
            return redirect('shop:product_list')
        
        for item in cart_items:
            item.total_price = item.product.price * item.quantity
        
        total_price = sum(item.product.price * item.quantity for item in cart_items)
        
        if request.method == 'POST':
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            email = request.POST.get('email')
            phone = request.POST.get('phone')
            address = request.POST.get('address')
            comments = request.POST.get('comments')
            
            order = Order.objects.create(
                user=request.user,
                first_name=first_name,
                last_name=last_name,
                email=email,
                phone=phone,
                address=address,
                comments=comments,
                total_price=total_price,
            )
            
            for item in cart_items:
                OrderItem.objects.create(
                    order=order,
                    product=item.product,
                    quantity=item.quantity,
                    price=item.product.price,
                )
            
            excel_file = generate_excel_receipt(order, cart_items)
            send_order_email(order, excel_file)
            cart_items.delete()
            
            messages.success(request, f'Заказ #{order.id} успешно оформлен! Чек отправлен на {email}')
            return redirect('shop:product_list')
        
        context = {
            'cart_items': cart_items,
            'total_price': total_price,
        }
        return render(request, 'shop/checkout.html', context)
    except Cart.DoesNotExist:
        messages.warning(request, 'Ваша корзина пуста.')
        return redirect('shop:product_list')


def generate_excel_receipt(order, cart_items):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Чек заказа"
    
    ws['A1'] = 'ЧЕК ЗАКАЗА'
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=16)
    ws.merge_cells('A1:C1')
    
    ws['A3'] = f'Номер заказа: #{order.id}'
    ws['A4'] = f'Дата: {order.created_at.strftime("%d.%m.%Y %H:%M")}'
    ws['A5'] = f'Покупатель: {order.first_name} {order.last_name}'
    ws['A6'] = f'Email: {order.email}'
    ws['A7'] = f'Телефон: {order.phone}'
    ws['A8'] = f'Адрес: {order.address}'
    
    ws['A10'] = 'Товар'
    ws['B10'] = 'Количество'
    ws['C10'] = 'Цена (руб.)'
    ws['D10'] = 'Сумма (руб.)'
    
    for row in ws['A10:D10']:
        for cell in row:
            cell.font = openpyxl.styles.Font(bold=True)
    
    row_num = 11
    for item in cart_items:
        ws[f'A{row_num}'] = item.product.name
        ws[f'B{row_num}'] = item.quantity
        ws[f'C{row_num}'] = float(item.product.price)
        ws[f'D{row_num}'] = float(item.product.price * item.quantity)
        row_num += 1
    
    ws[f'A{row_num}'] = 'ИТОГО:'
    ws[f'A{row_num}'].font = openpyxl.styles.Font(bold=True)
    ws[f'D{row_num}'] = float(order.total_price)
    ws[f'D{row_num}'].font = openpyxl.styles.Font(bold=True)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def send_order_email(order, excel_file):
    subject = f'Чек заказа #{order.id}'
    body = f'''
    Здравствуйте, {order.first_name} {order.last_name}!
    
    Ваш заказ #{order.id} успешно оформлен.
    
    Дата заказа: {order.created_at.strftime("%d.%m.%Y %H:%M")}
    Общая сумма: {order.total_price} руб.
    
    Адрес доставки: {order.address}
    
    Спасибо за покупку!
    '''
    email = EmailMessage(
        subject=subject,
        body=body,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[order.email],
    )
    email.attach(
        f'order_{order.id}_receipt.xlsx',
        excel_file.getvalue(),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    email.send()


def profile_page(request):
    return render(request, 'shop/profile.html')


class IsAdminOrReadOnly(BasePermission):
    def has_permission(self, request, view):
        if request.method in SAFE_METHODS:
            return True
        return (
            request.user.is_authenticated and 
            hasattr(request.user, 'profile') and 
            request.user.profile.role == 'ADMIN'
        )


@api_view(['POST'])
@permission_classes([AllowAny])
def api_register(request):
    username = request.data.get('username')
    password = request.data.get('password')
    email = request.data.get('email', '')
    
    if not username or not password:
        return Response({'error': 'Необходимо указать имя пользователя и пароль.'}, status=status.HTTP_400_BAD_REQUEST)
        
    if User.objects.filter(username=username).exists():
        return Response({'error': 'Пользователь с таким именем уже существует.'}, status=status.HTTP_400_BAD_REQUEST)
        
    user = User.objects.create_user(username=username, password=password, email=email)
    user.profile.full_name = username
    user.profile.save()
    
    login(request, user)
    return Response({'message': 'Вы успешно зарегистрировались.'}, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([AllowAny])
def api_login(request):
    username = request.data.get('username')
    password = request.data.get('password')
    
    user = authenticate(username=username, password=password)
    if user is not None:
        login(request, user)
        return Response({'message': 'Вход выполнен успешно.'}, status=status.HTTP_200_OK)
    return Response({'error': 'Неверный логин или пароль.'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_logout(request):
    logout(request)
    return Response({'message': 'Сессия успешно завершена.'}, status=status.HTTP_200_OK)


@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated])
def api_me(request):
    profile, created = Profile.objects.get_or_create(user=request.user)
    
    if request.method == 'GET':
        serializer = ProfileSerializer(profile)
        return Response(serializer.data)
        
    elif request.method == 'PATCH':
        serializer = ProfileSerializer(profile, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CategoryViewSet(ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer


class ProizvodstvoViewSet(ModelViewSet):
    queryset = Proizvodstvo.objects.all()
    serializer_class = ProizvodstvoSerializer


class ProductViewSet(ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsAdminOrReadOnly]


class CartViewSet(ModelViewSet):
    queryset = Cart.objects.all()
    serializer_class = CartSerializer

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class CartItemViewSet(ModelViewSet):
    queryset = CartItem.objects.all()
    serializer_class = CartItemSerializer


class OrderViewSet(ModelViewSet):
    serializer_class = OrderSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if hasattr(user, 'profile') and user.profile.role in ['ADMIN', 'MANAGER']:
            return Order.objects.all()
        return Order.objects.filter(user=user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    
